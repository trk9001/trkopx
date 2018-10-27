"""Tests for the FillSheet class' non-constructor methods."""

import os
import shutil

import pytest

from fillsheet import FillSheet
from openpyxl import load_workbook

# Assuming here that pytest is run from the root dir
file_name = 'tests/assets/Descr.xlsx'
dummy_file = 'tests/assets/DummyDescr.xlsx'


def test_get_number_of_rows():
    assert FillSheet.get_number_of_rows(file_name) == 15


def test_get_manufacturer_column_index():
    assert FillSheet.get_manufacturer_column_index(file_name) == 4


def test_config_rows_1():
    fs = FillSheet(file_name)
    fs.config_rows('3:10')
    assert fs.rows.start == 3 and fs.rows.end == 10


def test_config_rows_2():
    fs = FillSheet(file_name)
    fs.config_rows('3:20')
    assert fs.rows.start == 3 and fs.rows.end == 15


def test_config_rows_3():
    fs = FillSheet(file_name)
    fs.config_rows(':10')
    assert fs.rows.start == 2 and fs.rows.end == 10


def test_config_rows_4():
    fs = FillSheet(file_name)
    fs.config_rows(':')
    assert fs.rows.start == 2 and fs.rows.end == 15


def test_config_rows_start_gt_end_exception():
    with pytest.raises(ValueError):
        fs = FillSheet(file_name)
        fs.config_rows('10:3')


def test_config_rows_bad_format_exception():
    with pytest.raises(TypeError):
        fs = FillSheet(file_name)
        fs.config_rows([3, 10])


def test_half_fill():
    shutil.copy(file_name, dummy_file)
    fs = FillSheet(dummy_file, ':2')
    fs.half_fill()

    wb = load_workbook(dummy_file)
    ws = wb.active
    descr = ws.cell(2, 8).value
    os.remove(dummy_file)

    expected_descr = 'The {} from {} comes in {} colour, featuring'.format(
            'Limited Edition Trunks', 'Calvin Klein', 'Camo Print Black'
    )
    assert descr == expected_descr


def test_full_fill():
    shutil.copy(file_name, dummy_file)
    fs = FillSheet(dummy_file, ':3')
    fs.half_fill()

    wb = load_workbook(dummy_file)
    ws = wb.active
    ws.cell(2, 8).value += (' THESE FEATURES. This item also sports THESE '
                            'PROPERTIES.')
    ws.cell(3, 8).value += ' THESE FEATURES. Note: Blah blah.'
    ws.cell(3, 7).value = 'EZPZ'
    wb.save(dummy_file)

    fs.full_fill()

    wb = load_workbook(dummy_file)
    ws = wb.active
    descr1 = ws.cell(2, 7).value
    descr2 = ws.cell(3, 7).value
    os.remove(dummy_file)

    expected_descr1 = (('From {} comes the {} in {} colour, featuring THESE '
                        'PROPERTIES. This item also sports THESE FEATURES.')
                       .format('Calvin Klein', 'Limited Edition Trunks',
                               'Camo Print Black'))
    expected_descr2 = (('From {} comes the {} in {} colour, featuring THESE '
                        'FEATURES. Note: Blah blah.')
                       .format('Converse', 'CT All Star Hi Leather Trainers',
                               'Teak/Black/Driftwood'))
    assert descr1 == expected_descr1 and descr2 == expected_descr2
