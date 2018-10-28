"""The new fillsh.py (in development).

This script is intended to eventually replace fillsh.py, since that one hasn't
been updated in ages. Trying to reduce the number of *hacks* present in that
script. Please install `openpyxl` via pip before trying to use this script.
"""

import os.path
import re

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


class FillSheet:

    DEFAULT_FILE = 'Descr.xlsx'

    def __init__(self, file=DEFAULT_FILE, rows=None):
        """Set up all the required variables.

        Args:
            file -- Name of the Excel file
            rows -- Range of rows to work on

        Raises:
            FileNotFoundError
            TypeError
            ValueError
        """

        self.file = None
        self.rows = None
        self.seed = None
        self.max_rows = None

        # Validate and set the file name
        if '.xlsx' not in file:
            raise TypeError('INVALID FILE: '
                            'Must be a valid Excel file ending in .xlsx')
        elif not os.path.exists(file):
            raise FileNotFoundError('FILE DOES NOT EXIST')
        else:
            self.file = file

        # A struct for the `self.rows` data
        class Rows:
            def __init__(self):
                self.start = None
                self.end = None

        self.max_rows = self.get_number_of_rows(self.file)

        # Set the range of rows
        self.rows = Rows()
        if rows is None:
            self.rows.start = 2
            self.rows.end = self.max_rows
        else:
            self.config_rows(rows)

        # Set the seed column index
        self.seed = self.get_manufacturer_column_index(self.file)
        if self.seed is None:
            raise ValueError('MANUFACTURER COLUMN\'S INDEX NOT FOUND')

    def config_rows(self, rows):
        """Set the range of rows from a string."""

        if isinstance(rows, str) and re.match(r'^\d*:\d*$', rows):
            row_values = rows.split(':')
            if row_values[0]:
                self.rows.start = int(row_values[0])
            else:
                self.rows.start = 2
            if row_values[1]:
                self.rows.end = int(row_values[1])
                if self.rows.end > self.max_rows:
                    self.rows.end = self.max_rows
            else:
                self.rows.end = self.max_rows

            if self.rows.start > self.rows.end:
                raise ValueError('INVALID VALUE FOR ROWS: '
                                 'START must be less than END')
        else:
            raise TypeError('INVALID TYPE FOR ROWS: '
                            'Must be empty or of the form "START:END"')

    @staticmethod
    def get_number_of_rows(file):
        """Return the number of rows in a worksheet."""
        wb = load_workbook(file)
        return len(list(wb.active.rows))

    @staticmethod
    def get_manufacturer_column_index(file):
        """Return the index of the Manufacturer column in a worksheet."""
        wb = load_workbook(file)
        ws = wb.active

        # The Manufacturer column is usually among the first ones,
        # barring the very first column.
        for col in range(2, 10):
            if ws.cell(1, col).value == 'MANUFACTURER':
                return col

    def half_fill(self):
        """Generate boilerplate text in the second description column."""

        wb = load_workbook(self.file)
        ws = wb.active

        # Set column numbers sequentially from the seed value, for columns:
        # manufacturer, product, colour, fulldescr1 and fulldescr2.
        mnf, pdt, clr, dc1, dc2 = (self.seed + i for i in range(5))

        i = self.rows.start - 1

        while i < self.rows.end:
            i += 1

            # If there is any text in the second description column, move
            # it to the first description column.
            if ws.cell(i, dc2).value:
                tmp = ws.cell(i, dc1).value
                if tmp is None:
                    tmp = ''
                ws.cell(i, dc1).value = tmp + '; ' + ws.cell(i, dc2).value
                ws.cell(i, dc2).value = None

            # Skip consecutive duplicates of the last product
            if ws.cell(i, pdt).value == ws.cell(i - 1, pdt).value:
                continue

            # Construct a description from the columns' data, format the cell
            # and save it.
            descr = 'The {} from {} comes in {} colour, featuring'.format(
                    *[ws.cell(i, col).value for col in [pdt, mnf, clr]]
            )
            self.format_cell(ws.cell(i, dc2))
            ws.cell(i, dc2).value = descr

        wb.save(self.file)

    def full_fill(self):
        """Generate full descriptions in the first column."""

        wb = load_workbook(self.file)
        ws = wb.active

        # Compiled regex objects
        rgx1 = re.compile(r'featuring.*$')
        rgx2 = re.compile(r'featuring ([^.]*)[.]')
        rgx3 = re.compile(r'sports? (.*)[.]$')

        # Manual control codes
        code_skip = 'SKIP'
        code_skip_reset = 'RESET'
        code_skip_note = 'EZPZ'

        # Set column numbers sequentially from the seed value, for columns:
        # manufacturer, product, colour, fulldescr1 and fulldescr2.
        mnf, pdt, clr, dc1, dc2 = (self.seed + i for i in range(5))

        # Reset variables used for repetition checking
        mnf_val = pdt_val = clr_val = None
        times_repeated = 0

        i = self.rows.start - 1

        while i < self.rows.end:
            i += 1

            # Manual skip condition
            if ((ws.cell(i, dc1).value
                 and code_skip in ws.cell(i, dc1).value)
                    or (ws.cell(i, dc2).value
                        and code_skip in ws.cell(i, dc2).value)):
                continue

            # Manual recurrence reset condition
            if ((ws.cell(i, dc1).value
                 and code_skip_reset in ws.cell(i, dc1).value)
                    or (ws.cell(i, dc2).value
                        and code_skip_reset in ws.cell(i, dc2).value)):
                pdt_val = None
                continue

            # If the current product has been repeated consecutively
            if ws.cell(i, pdt).value == pdt_val:
                new_clr_val = ws.cell(i, clr).value
                if times_repeated == 0 or times_repeated == 2:
                    new_dc1 = (ws.cell(i - 1, dc2).value
                               .replace(clr_val, new_clr_val))
                    new_dc2 = (ws.cell(i - 1, dc1).value
                               .replace(clr_val, new_clr_val))
                elif times_repeated == 1:
                    new_dc1 = (ws.cell(i - 2, dc1).value
                               .replace('From {} comes'.format(mnf_val),
                                        '{} offers'.format(mnf_val))
                               .replace(clr_val, new_clr_val))
                    new_dc2 = (ws.cell(i - 2, dc2).value
                               .replace('The {} from {}'
                                        .format(pdt_val, mnf_val),
                                        'Offered by {}, the {}'
                                        .format(mnf_val, pdt_val))
                               .replace(clr_val, new_clr_val))
                elif times_repeated == 3:
                    new_dc1 = (ws.cell(i - 4, dc1).value
                               .replace(clr_val, new_clr_val))
                    new_dc2 = (ws.cell(i - 4, dc2).value
                               .replace(clr_val, new_clr_val))
                    # Repeat the cycle in case of more recurrences
                    times_repeated = -1
                else:
                    # Probably not going to reach here
                    continue

                self.format_cell(ws.cell(i, dc1))
                ws.cell(i, dc1).value = new_dc1
                self.format_cell(ws.cell(i, dc2))
                ws.cell(i, dc2).value = new_dc2

                times_repeated += 1

            # Regular case (non-repeated)
            else:
                times_repeated = 0

                # Necessary to check for repetitions
                mnf_val = ws.cell(i, mnf).value
                pdt_val = ws.cell(i, pdt).value
                clr_val = ws.cell(i, clr).value

                # Typical second description generation
                part1 = 'From {} comes the {} in {} colour, '.format(
                        mnf_val, pdt_val, clr_val
                )
                try:
                    part2 = rgx1.search(ws.cell(i, dc2).value).group(0)
                except AttributeError:
                    raise ValueError('"FEATURING" CLAUSE NOT FOUND: '
                                     'In row {}'.format(i))
                descr = part1 + part2

                # Modifications in absence of ending note
                if ws.cell(i, dc1).value == code_skip_note:
                    pass
                else:
                    try:
                        featuring_clause = rgx2.search(descr).group(1)
                        sports_clause = rgx3.search(descr).group(1)
                    except AttributeError:
                        raise ValueError('CLAUSE NOT FOUND: '
                                         'In row {}'.format(i))
                    except IndexError:
                        raise ValueError('SUB-CLAUSE EMPTY: '
                                         'In row {}'.format(i))

                    descr = descr.replace(sports_clause, '!TEMP!')
                    descr = descr.replace(featuring_clause, sports_clause)
                    descr = descr.replace('!TEMP!', featuring_clause)

                self.format_cell(ws.cell(i, dc1))
                ws.cell(i, dc1).value = descr

        wb.save(self.file)

    @staticmethod
    def format_cell(cell):
        """Formats a cell from a worksheet opened using openpyxl."""
        cell.alignment = Alignment(horizontal='left', wrap_text=True)
        cell.font = Font(name='Calibri', size=8)
